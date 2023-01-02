# !/usr/bin/env python3
# -*- coding:utf-8 -*-

import re
import os
import bs4
import openpyxl
from openpyxl.utils import get_column_letter
import urllib.request
import urllib.error
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import pymysql
from pymysql.converters import escape_string
from sqlalchemy import create_engine
import pandas as pd
from config import config


problem_list = ['P', 'B', 'CF', 'SP', 'AT', 'UVA']
problem_list_name = ['主题库', '入门题库', 'CF题库', 'SPOJ题库', 'AtCoder题库', 'UVA题库']


def write_mysql(data_list):
    connection = pymysql.connect(host=config['HOST'], port=config['PORT'], user=config['USERNAME'],
                                 password=config['PASSWORD'], charset='utf8', database='luogu_problems')
    cursor = connection.cursor()
    cursor.execute("USE luogu_problems")
    connection.commit()
    engine = create_engine(f"mysql+pymysql://{config['USERNAME']}:{config['PASSWORD']}"
                           f"@{config['HOST']}:{config['PORT']}/luogu_problems?charset=utf8")
    for data in data_list:
        sql = f"""INSERT INTO luogu_problems.problems VALUE("{escape_string(data[0])}","{escape_string(data[1])}","{escape_string(data[4])}");"""
        try:
            cursor.execute(sql)
        except pymysql.err.IntegrityError:
            continue
        connection.commit()
        if len(data[2]) != 0:
            for sour in data[2]:
                sql = f"SELECT id FROM tags WHERE name='{sour}'"
                df = pd.read_sql(sql, engine)
                tag_id = df.iat[0, 0]
                sql = f"INSERT INTO luogu_problems.tags_problems_link(problem_title, tag_id) VALUE('{data[0]}', '{int(tag_id)}');"
                cursor.execute(sql)
                connection.commit()
        if len(data[3]) != 0:
            for algo in data[3]:
                sql = f"SELECT id FROM tags WHERE name='{algo}'"
                df = pd.read_sql(sql, engine)
                tag_id = df.iat[0, 0]
                sql = f"INSERT INTO luogu_problems.tags_problems_link(problem_title, tag_id) VALUE('{data[0]}', '{tag_id}');"
                cursor.execute(sql)
                connection.commit()


def set_width():
    wb = openpyxl.load_workbook('.\\list.xlsx')
    ws = wb['题目']
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                cell_len = 0.7*len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 8
    wb.save('.\\list.xlsx')


def write_excel(data_list):
    if os.path.exists(".\\list.xlsx"):
        workbook = openpyxl.load_workbook(".\\list.xlsx")
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet('题目')
        workbook.remove(workbook['Sheet'])
        title = ['题号', '题目名称', '来源', '算法', '难度']
        sheet.append(title)
    sheet = workbook['题目']
    for i in data_list:
        sheet.append(i)
    workbook.save(".\\list.xlsx")


def findn(n, s) -> bool:
    with open(".\\.done\\" + n + '.txt', "r") as e:
        for i in e:
            if str(i) == str(s)+'\n':
                return True
        return False


def get_md(html):
    bs = bs4.BeautifulSoup(html, "html.parser")
    core = bs.select("article")[0]
    md = str(core)
    md = re.sub("<h1>", "# ", md)
    md = re.sub("<h2>", "## ", md)
    md = re.sub("<h3>", "### ", md)
    md = re.sub("</?[a-zA-Z]+[^<>]*>", "", md)
    return md


def save_data(data, filename):
    cfilename = ".\\problems\\" + filename
    file = open(cfilename, "w", encoding="utf-8")
    for d in data:
        file.writelines(d)
    file.close()


def get_html(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36"
    }
    request = urllib.request.Request(url=url, headers=headers)
    try:
        response = urllib.request.urlopen(request)
    except urllib.error.URLError:
        print('当前网络不稳定，请稍后再试')
        return "internet"
    html = response.read().decode('utf-8')
    if str(html).find("HttpException") == -1:
        return html
    else:
        return "error"


def problem(s) -> bool:
    base_url = "https://www.luogu.com.cn/problem/"
    if os.path.exists(".\\problems\\" + s + '.md'):
        return True
    print("正在爬取{}...".format(s), end="")
    html = get_html(base_url + s)
    if html == "internet":
        return False
    if html == "error":
        print("爬取失败，可能是无权查看")
        return True
    problem_md = get_md(html)
    print("爬取成功！正在保存...", end="")
    save_data(problem_md, s + ".md")
    print("保存成功！")
    return True


def problems(problem_type) -> bool:
    browser = webdriver.Chrome()
    url = r"https://www.luogu.com.cn/problem/list?type=" + problem_type + '&page='
    try:
        browser.get(url+"1")
    except selenium.common.exceptions.WebDriverException:
        print("网络不稳定，请稍后再试")
        browser.quit()
        return False
    WebDriverWait(browser, 1000).until(
        ec.presence_of_all_elements_located((By.XPATH, r"/html/body/div/div[2]/main/div/div/div/div[2]/div/div/span/strong"))
    )
    pages = int(browser.find_element(By.XPATH, r'/html/body/div/div[2]/main/div/div/div/div[2]/div/div/span/strong').text)
    print("一共要检查{}页".format(pages))
    for i in range(1, pages+1):
        data_list1 = []
        data_list2 = []
        print("正在检查第{}页".format(i))
        try:
            browser.get(url+str(i))
        except selenium.common.exceptions.WebDriverException:
            print("网络不稳定，请稍后再试")
            browser.quit()
            return False
        WebDriverWait(browser, 1000).until(
            ec.presence_of_all_elements_located(
                (By.XPATH, r"/html/body/div/div[2]/main/div/div/div/div[2]/div/div/span/strong"))
        )
        divs = browser.find_elements(By.XPATH, '/html/body/div/div[2]/main/div/div/div/div[1]/div[2]/div')
        numb_list = []
        title_list = []
        sour_list = []
        algo_list = []
        sour_str_list = []
        algo_str_list = []
        difficulty_list = []
        flag = False
        for div in divs:
            numb = div.find_element(By.XPATH, 'span[2]').text
            if os.path.exists(".\\problems\\"+numb+'.md'):
                continue
            else:
                flag = True
            title = div.find_element(By.XPATH, 'div[1]/a').text
            temp = div.find_elements(By.XPATH, 'div[2]/div/a')
            sour = []
            for tem in temp:
                sour.append(tem.text)
            sour_str = '，'.join(sour)
            difficulty = div.find_element(By.XPATH, 'div[3]/a/span').text
            numb_list.append(numb)
            title_list.append(title)
            sour_str_list.append(sour_str)
            sour_list.append(sour)
            difficulty_list.append(difficulty)
            if not problem(str(numb)):
                browser.quit()
                return False
        if flag:
            browser.find_element(By.XPATH, '/html/body/div/div[2]/main/div/div/div/div[1]/div[1]/div/div[4]/span/a').click()
            for div in divs:
                temp = div.find_elements(By.XPATH, 'div[2]/div/a')
                algo = []
                for tem in temp:
                    algo.append(tem.text)
                algo_str = '，'.join(algo)
                algo_str_list.append(algo_str)
                algo_list.append(algo)
            for numb, title, sour, algo, difficulty in zip(numb_list, title_list, sour_str_list, algo_str_list,
                                                           difficulty_list):
                data = [numb, title, sour, algo, difficulty]
                data_list1.append(data)
            for numb, title, sour, algo, difficulty in zip(numb_list, title_list, sour_list, algo_list, difficulty_list):
                data = [numb, title, sour, algo, difficulty]
                data_list2.append(data)
            if config['USE_MYSQL']:
                write_mysql(data_list2)
            write_excel(data_list1)
        print("第{}页检查完毕".format(i))
    browser.quit()
    return True


def main():
    print("即将开始检查")
    for i, j in zip(problem_list, problem_list_name):
        print("准备爬检查"+j)
        if not problems(i):
            return
        print(j+"检查完毕")


if __name__ == '__main__':
    main()
