#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import bs4
import re
import os
import openpyxl
from openpyxl.utils import get_column_letter


def set_width():
    wb = openpyxl.load_workbook('.\\list.xlsx')
    ws = wb['题目']
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                cell_len = 0.7*len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 8
    wb.save('.\\list.xlsx')


def write_excel(data_list):
    if os.path.exists(".\\list.xlsx"):
        workbook = openpyxl.load_workbook(".\\list.xlsx")
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet('题目')
        workbook.remove(workbook['Sheet'])
        title = ['题号', '题目名称', '来源', '算法', '难度']
        sheet.append(title)
    sheet = workbook['题目']
    for i in data_list:
        sheet.append(i)
    workbook.save(".\\list.xlsx")
    set_width()


def findn(n, s) -> bool:
    with open(".\\.done\\" + n + '.txt', "r") as e:
        for i in e:
            if str(i) == str(s)+'\n':
                return True
        return False


def getMD(html):
    bs = bs4.BeautifulSoup(html, "html.parser")
    core = bs.select("article")[0]
    md = str(core)
    md = re.sub("<h1>", "# ", md)
    md = re.sub("<h2>", "## ", md)
    md = re.sub("<h3>", "### ", md)
    md = re.sub("</?[a-zA-Z]+[^<>]*>", "", md)
    return md


def saveData(data, filename):
    cfilename = ".\\problems\\" + filename
    file = open(cfilename, "w", encoding="utf-8")
    for d in data:
        file.writelines(d)
    file.close()
